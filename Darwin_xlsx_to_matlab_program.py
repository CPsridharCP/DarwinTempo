# Cleans up huge bag files extracted from ROS in csv to .mat files
# Simulation of Darwin - Walking
# Coded by CP Sridhar

import openpyxl 
import scipy.io

wb = openpyxl.Workbook()
wb = openpyxl.load_workbook(filename = 'darwin_joint_states.xlsx')
#wb = openpyxl.load_workbook(filename = 'ten_row_sampleset.xlsx')

# To display all of the available worksheet names
sheets =  wb.sheetnames
print sheets

# Selecting the first Sheet
ws = wb[sheets[0]]

# Finding the number of rows
row_count = ws.max_row
print ("Row Count: " + str(row_count))

# Finding Joint Names and adding them to joint_names array
joint_names = ws['H2'].value
joint_names = joint_names.replace("[","")
joint_names = joint_names.replace("]","")
joint_names = joint_names.replace("'","")
joint_names = joint_names.replace("j_","")
joint_names = [str(names) for names in joint_names.split(",")]

#Creating empty lists to hold all our data - Since we have 24 jonits this will be a list of 24 sub lists
# in python we call arrays as lists
total_position =[[],[],[],[],[],[],[],[],[],[],[],[],[],[],[],[],[],[],[],[],[],[],[],[]]
total_velocity =[[],[],[],[],[],[],[],[],[],[],[],[],[],[],[],[],[],[],[],[],[],[],[],[]]
total_torque   =[[],[],[],[],[],[],[],[],[],[],[],[],[],[],[],[],[],[],[],[],[],[],[],[]]
total_time = []

# Next line is only to confirm we created 24 slots
# print ("total_size: "+str(len(total_position)))

# Iterating through each row in the excel sheet
for i in range(2,row_count+1):

	#POSITION
	position_cell = "I"+str(i)
	position = ws[position_cell].value
	position = position.replace("[","")
	position = position.replace("]","")
	position = position.replace("u'","")
	position = position.replace("'","")
	position = [value for value in position.split(",")]
	for j in range(0,24):
		total_position[j].append(float(position[j]))

	#VELOCITY
	velocity_cell = "J"+str(i)
	velocity = ws[velocity_cell].value
	velocity = velocity.replace("[","")
	velocity = velocity.replace("]","")
	velocity = velocity.replace("u'","")
	velocity = velocity.replace("'","")
	velocity = [value for value in velocity.split(",")]
	for j in range(0,24):
		total_velocity[j].append(float(velocity[j]))

    #TORQUE
	torque_cell = "K"+str(i)
	torque = ws[torque_cell].value
	torque = torque.replace("[","")
	torque = torque.replace("]","")
	torque = torque.replace("u'","")
	torque = torque.replace("'","")
	torque = [value for value in torque.split(",")]
	for j in range(0,24):
		total_torque[j].append(float(torque[j]))

    #TIME LINE
	time_sec_cell = "E"+str(i)
	time_nano_sec_cell = "F"+str(i)
	timeS = ws[time_sec_cell].value
	timeNS = ws[time_nano_sec_cell].value
	final_number_string = str(timeS)+"."+str(timeNS)
	total_time.append(float(final_number_string))

# this saves our arrays in matlab readable format
# SYNTAX = scipy.io.savemat('<Directort/File_Name.mat>', mdict={'arr': <array_name>}) 
scipy.io.savemat('/Users/cplabs/Documents/Projects/C_Mummolo/Output_Matlab/Position_Darwin.mat', mdict={'arr': total_position})
scipy.io.savemat('/Users/cplabs/Documents/Projects/C_Mummolo/Output_Matlab/Velocity_Darwin.mat', mdict={'arr': total_velocity})
scipy.io.savemat('/Users/cplabs/Documents/Projects/C_Mummolo/Output_Matlab/Torque_Darwin.mat', mdict={'arr': total_torque})
scipy.io.savemat('/Users/cplabs/Documents/Projects/C_Mummolo/Output_Matlab/TimeStamp_Darwin.mat', mdict={'arr': total_time})
scipy.io.savemat('/Users/cplabs/Documents/Projects/C_Mummolo/Output_Matlab/JointNames_Darwin.mat', mdict={'arr': joint_names})

""" TO PRINT OUT ALL DATA UNCOMMENT THE FOLOWING 13 LINES
print joint_names
print("")
print ("TOTAL Position:")
print total_position
print("")
print ("TOTAL Velocity:")
print total_velocity
print("")
print ("TOTAL Torque:")
print total_torque
print("")
print ("TOTAL Time:")
print total_time
"""


